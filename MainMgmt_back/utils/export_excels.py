import pandas as pd
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
import logging
from urllib.parse import quote

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Excel导出工具类
    支持两种数据源：
    1. Django QuerySet - 自动从模型字段提取数据
    2. 纯列表/元组数据 - 直接使用预构建的数据行
    """

    def __init__(self, queryset, filename, table_name, headers):
        """
        初始化导出器
        :param queryset: Django QuerySet 或 list/tuple 类型的数据行
        :param filename: 导出文件名（支持 f-string 格式）
        :param table_name: Excel工作表名称
        :param headers: 表头配置列表，如 [{'titlename': '列名', 'col_width': 20}]
        """
        self.queryset = queryset
        self.filename = str(filename)
        self.table_name = table_name
        self.headers = headers
        logger.info(f"ExcelExporter 初始化 - 文件名: '{self.filename}'")

    def export(self):
        """执行导出操作，返回HttpResponse"""
        # 创建HTTP响应
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # ✅ 核心修复：使用 RFC 5987 标准的 filename* 参数
        # 对中文文件名进行 UTF-8 编码，适配所有现代浏览器
        encoded_filename = quote(self.filename.encode('utf-8'))
        response[
            'Content-Disposition'] = f"attachment; filename*=utf-8''{encoded_filename}; filename=\"{self.filename}\""

        logger.info(f"开始导出Excel - 文件名: {self.filename}, 工作表名: {self.table_name}")

        # 使用Pandas将DataFrame写入Excel（通过BytesIO）
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 判断是QuerySet还是纯数据列表
            if isinstance(self.queryset, (list, tuple)):
                # 纯数据列表的情况
                df_columns = [header['titlename'] for header in self.headers]
                if not self.queryset:
                    df = pd.DataFrame(columns=df_columns)
                    logger.warning("导出数据为空，创建空DataFrame")
                else:
                    df = pd.DataFrame(self.queryset, columns=df_columns)
                    logger.info(f"从列表创建DataFrame，共 {len(self.queryset)} 行")
            else:
                # 原始的QuerySet情况
                field_to_title = {field.name: header['titlename'] for field, header in
                                  zip(self.queryset.model._meta.get_fields(include_parents=False), self.headers)}
                df_data = list(self.queryset.values_list(*field_to_title.keys()))
                df = pd.DataFrame(df_data, columns=field_to_title.values())
                logger.info(f"从QuerySet创建DataFrame，共 {len(df_data)} 行")

            # 设置字体和颜色
            font_color = Color(rgb='000000')
            yichang_font_color = Color(rgb='FF0000')  # 红色
            font_style2 = Font(name='微软雅黑', sz=12, b=False, color=font_color)
            font_style3 = Font(name='h1', sz=11, b=False, color=yichang_font_color)

            df.to_excel(writer, sheet_name=self.table_name, index=False, startrow=1)
            workbook = writer.book
            worksheet = writer.sheets[self.table_name]

            # 设置纸张大小和方向
            worksheet.page_setup.paper_size = 8  # 8=A3纸,9=A4
            worksheet.page_setup.orientation = 'landscape'  # 横向

            # 设置列宽
            for col_idx, record in enumerate(self.headers, start=1):
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = record['col_width']

            # 设置外边框为细实线
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            merge_range = f'A1:{get_column_letter(len(self.headers))}1'
            worksheet.merge_cells(merge_range)
            worksheet['A1'].font = Font(name='微软雅黑', sz=16, b=True, color=Color(rgb='000000'))
            worksheet.cell(1, 1).value = self.table_name
            worksheet.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

            # 冻结首行
            worksheet.freeze_panes = 'A2'
            align = Alignment(horizontal='center', vertical='center', wrapText=True)

            # 两层循环遍历所有有数据的单元格，对每次单元格进行样式设置
            for i in range(1, worksheet.max_row + 1):
                for j in range(1, worksheet.max_column + 1):
                    worksheet.cell(i, j).border = thin_border
                    worksheet.cell(i, j).alignment = align
                    # 定义"异常"的字体颜色
                    if (str(worksheet.cell(i, j).value).strip() == '坏卡'):
                        worksheet.cell(i, j).font = font_style3
                    if (i == 2):
                        worksheet.cell(2, j).font = font_style2  # 定义表格标题字体和颜色

        # 保存Excel到response
        output.seek(0)
        response.write(output.getvalue())
        logger.info(f"Excel导出成功，文件大小: {len(output.getvalue())} bytes")

        return response